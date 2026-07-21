"""
Build the Landed Cost & Margin Scenario workbook.

Five tabs, all formula-driven off Tab 1 (Assumptions):
  1. Assumptions          - every input/driver in one place (blue = editable)
  2. Landed Cost Impact   - per-SKU landed cost, air vs sea, FX effect vs freight effect
  3. Revenue Scenario     - monthly revenue by segment, elasticity-driven, base/mod/severe + variance
  4. Margin Bridge        - waterfall from planned GP to scenario GP (FX / air / sea sub / demand)
  5. Mitigation Levers    - shift-to-sea / hedge / price passthrough, $ impact each, gap-closing

Run:  python scripts/build_landed_cost_scenarios.py
Then: python /root/.claude/skills/xlsx/scripts/recalc.py Landed_Cost_Margin_Scenarios.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, Series

OUT = "Landed_Cost_Margin_Scenarios.xlsx"

# ---- shared style constants ----------------------------------------------
FONT = "Arial"
NAVY = "1F3864"
LIGHTBLUE = "D9E1F2"
YELLOW = "FFF2CC"
BLUE = "0000FF"      # hardcoded inputs
GREEN = "006100"     # pure cross-sheet links
BLACK = "000000"

CUR0 = '$#,##0;($#,##0);"-"'          # whole AUD
CUR2 = '$#,##0.00;($#,##0.00);"-"'    # per-unit AUD (cents matter)
PCT1 = '0.0%;(0.0%);"-"'
PCT1S = '+0.0%;-0.0%;0.0%'            # signed % (for shifts vs plan)
RATE3 = '0.0000'
NUM0 = '#,##0'
MULT = '0.000x'

thin = Side(style="thin", color="BFBFBF")
med = Side(style="medium", color=NAVY)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_BORDER = Border(bottom=med)

AS = "'1. Assumptions'!"      # assumptions sheet reference prefix
RS = "'3. Revenue Scenario'!"
MB = "'4. Margin Bridge'!"


def w(ws, coord, value=None, *, bold=False, color=BLACK, size=10, italic=False,
      fmt=None, fill=None, align=None, wrap=False, border=None, valign=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, bold=bold, color=color, size=size, italic=italic)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align or valign:
        c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    elif wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    if border:
        c.border = border
    return c


def title(ws, coord, text, span):
    w(ws, coord, text, bold=True, color="FFFFFF", size=14, fill=NAVY,
      align="left", valign="center")
    ws.merge_cells(f"{coord}:{span}")
    ws.row_dimensions[int(coord[1:])].height = 24


def section(ws, row, text, last_col="H"):
    w(ws, f"A{row}", text, bold=True, color="FFFFFF", size=11, fill=NAVY, valign="center")
    ws.merge_cells(f"A{row}:{last_col}{row}")
    ws.row_dimensions[row].height = 18


def colhdr(ws, row, headers, start=1):
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers):
        col = get_column_letter(start + i)
        w(ws, f"{col}{row}", h, bold=True, size=10, fill=LIGHTBLUE,
          align="center" if i else "left", wrap=True, valign="center", border=HDR_BORDER)
    ws.row_dimensions[row].height = 28


wb = openpyxl.Workbook()

# =========================================================================
# TAB 1 - ASSUMPTIONS & DRIVERS
# =========================================================================
s1 = wb.active
s1.title = "1. Assumptions"
s1.sheet_view.showGridLines = False
title(s1, "A1", "Landed Cost & Margin Scenario Model  —  Assumptions & Drivers", "G1")
w(s1, "A2", "Planning horizon: Sep 2026 – Jun 2027.  BLUE cells are inputs — edit them and every tab recalculates. "
             "Base / Moderate / Severe are downside stress cases (Base = mild pressure, Severe = worst case).",
  italic=True, size=9, color="595959", wrap=True)
s1.merge_cells("A2:G2")
s1.row_dimensions[2].height = 26

# --- scenario drivers ---
section(s1, 4, "SCENARIO DRIVERS", "G")
colhdr(s1, 5, ["Driver", "Plan", "Base", "Moderate", "Severe", "Unit", "Notes"])

driver_rows = [
    # label, plan, base, mod, sev, unit, note, fmt
    (6,  "AUD/USD spot rate",        0.660, 0.655, 0.620, 0.580, "USD per 1 AUD", "Lower = weaker AUD = dearer imports", RATE3),
    (8,  "Air freight rate",         6.00,  6.20,  7.50,  9.00,  "USD per kg",    "Spot air out of origin", CUR2),
    (9,  "Sea freight rate",         180,   190,   260,   340,   "USD per CBM",   "FCL-equivalent per cubic metre", CUR0),
    (10, "Import duty",              0.05,  0.05,  0.05,  0.05,  "% of FOB",      "Held flat across scenarios", PCT1),
    (11, "Insurance & handling",     0.03,  0.03,  0.03,  0.03,  "% of FOB",      "Held flat across scenarios", PCT1),
    (12, "Consumer spending vs plan",0.00, -0.02, -0.05, -0.11, "% vs plan",     "Macro driver of demand", PCT1S),
    (13, "Air->sea substitution shift",0.00,0.00, 0.10,  0.20,  "pp of units",   "Defensive shift embedded in scenario", PCT1S),
]
for row, label, pl, ba, mo, se, unit, note, fmt in driver_rows:
    w(s1, f"A{row}", label, size=10)
    for col, val in zip("BCDE", (pl, ba, mo, se)):
        w(s1, f"{col}{row}", val, color=BLUE, fmt=fmt, align="center")
    w(s1, f"F{row}", unit, size=9, color="595959")
    w(s1, f"G{row}", note, size=9, color="595959")

# derived AUD per USD row 7 (formula, black)
w(s1, "A7", "AUD per USD  (= 1 / spot)", size=10, italic=True)
for col in "BCDE":
    w(s1, f"{col}7", f"=1/{col}6", fmt=RATE3, align="center")
w(s1, "F7", "AUD per 1 USD", size=9, color="595959")
w(s1, "G7", "Derived. Every USD cost is multiplied by this.", size=9, italic=True, color="595959")

# --- elasticity by segment ---
section(s1, 15, "DEMAND ELASTICITY BY SEGMENT", "G")
colhdr(s1, 16, ["Segment", "Elasticity", "", "", "", "", "Notes"])
elast = [(17, "Premium", 1.5, "Discretionary — most sensitive to spending"),
         (18, "Core",    1.0, "Moves roughly with the market"),
         (19, "Value",   0.5, "Defensive — customers trade down into it")]
for row, seg, e, note in elast:
    w(s1, f"A{row}", seg)
    w(s1, f"B{row}", e, color=BLUE, fmt="0.0", align="center")
    w(s1, f"G{row}", note, size=9, color="595959")
w(s1, "C16", "demand % move per 1% spending move", size=8, italic=True, color="595959")
s1.merge_cells("C16:F16")

# --- aggregate model basis ---
section(s1, 21, "AGGREGATE MODEL BASIS  (Margin Bridge & Levers)", "G")
colhdr(s1, 22, ["Item", "Value", "", "", "", "Unit", "Notes"])
basis = [
    (23, "Total planned units (period)", 480000, NUM0, "units",       "Representative volume, Sep26-Jun27"),
    (24, "Planned ASP (avg selling price)", 38.00, CUR2, "AUD/unit",  "Blended across catalogue"),
    (25, "Avg FOB cost", 12.00, CUR2, "USD/unit",  "Representative average (cf. Tab 2 SKUs)"),
    (26, "Avg weight", 0.90, "0.00", "kg/unit",     "Drives air freight per unit"),
    (27, "Avg units per CBM", 220, NUM0, "units/CBM","Drives sea freight per unit"),
    (28, "Planned air share of units", 0.35, PCT1, "% of units", "Balance ships by sea"),
    (29, "Price elasticity", 0.80, "0.00", "vol% per 1% price", "For the price-passthrough lever"),
]
for row, label, val, fmt, unit, note in basis:
    w(s1, f"A{row}", label)
    w(s1, f"B{row}", val, color=BLUE, fmt=fmt, align="center")
    w(s1, f"F{row}", unit, size=9, color="595959")
    w(s1, f"G{row}", note, size=9, color="595959")

for col, wd in zip("ABCDEFG", (30, 11, 11, 11, 11, 15, 42)):
    s1.column_dimensions[col].width = wd

# =========================================================================
# TAB 2 - LANDED COST IMPACT
# =========================================================================
s2 = wb.create_sheet("2. Landed Cost Impact")
s2.sheet_view.showGridLines = False
title(s2, "A1", "Landed Cost Impact  —  Air vs Sea, FX effect vs Freight effect", "N1")
w(s2, "A2", "All landed costs are AUD per unit. Effects are measured vs Plan. Attribution order: FX first "
             "(revalue every USD cost at the scenario rate), then Freight (apply the scenario freight rate). "
             "FX Δ + Freight Δ = Total Δ, exactly.", italic=True, size=9, color="595959", wrap=True)
s2.merge_cells("A2:N2")
s2.row_dimensions[2].height = 30

# SKU master
section(s2, 4, "SKU MASTER  (inputs)", "N")
colhdr(s2, 5, ["SKU Code", "Description", "FOB Cost (USD)", "Weight (kg)", "Units / CBM"])
skus = [
    ("SKU-001", "Compact appliance",     24.00, 2.10, 60),
    ("SKU-002", "Kitchen accessory",      6.50, 0.40, 480),
    ("SKU-003", "Homeware set",          14.00, 1.80, 90),
    ("SKU-004", "Small electronic",      32.00, 0.60, 300),
    ("SKU-005", "Textile goods",          9.00, 1.20, 140),
    ("SKU-006", "Bulky furniture item",  45.00, 8.50, 12),
]
for i, (code, desc, fob, wt, cbm) in enumerate(skus):
    r = 6 + i
    w(s2, f"A{r}", code, bold=True)
    w(s2, f"B{r}", desc)
    w(s2, f"C{r}", fob, color=BLUE, fmt=CUR2, align="center")
    w(s2, f"D{r}", wt, color=BLUE, fmt="0.00", align="center")
    w(s2, f"E{r}", cbm, color=BLUE, fmt=NUM0, align="center")


def landed_expr(mode, m, fxcol, ratecol):
    """AUD landed cost per unit for SKU master-row m, at assumptions FX column fxcol / rate column ratecol."""
    prod = f"$C${m}*(1+{AS}$B$10+{AS}$B$11)"
    if mode == "air":
        fr = f"$D${m}*{AS}${ratecol}$8"
    else:
        fr = f"(1/$E${m})*{AS}${ratecol}$9"
    return f"{AS}${fxcol}$7*({prod}+{fr})"


SCEN = [("Base", "C"), ("Moderate", "D"), ("Severe", "E")]  # (label, assumptions column)


def landed_block(ws, start_row, mode, master_first):
    hdr = ["SKU", "Plan"]
    for label, _ in SCEN:
        hdr += [f"{label}\nLanded", f"{label}\nFX Δ", f"{label}\nFreight Δ", f"{label}\nTotal Δ"]
    colhdr(ws, start_row, hdr)
    from openpyxl.utils import get_column_letter
    for i in range(6):
        r = start_row + 1 + i
        m = master_first + i
        w(ws, f"A{r}", f"=$A${m}")               # SKU code linked from master
        w(ws, f"B{r}", "=" + landed_expr(mode, m, "B", "B"), fmt=CUR2)  # Plan
        for j, (label, xcol) in enumerate(SCEN):
            base_c = 3 + j * 4                    # first col of this scenario block (C=3)
            L = get_column_letter(base_c)
            FX = get_column_letter(base_c + 1)
            FR = get_column_letter(base_c + 2)
            TT = get_column_letter(base_c + 3)
            # Landed at scenario FX + scenario rate
            w(ws, f"{L}{r}", "=" + landed_expr(mode, m, xcol, xcol), fmt=CUR2)
            # FX effect: scenario FX, PLAN rate, minus Plan
            w(ws, f"{FX}{r}", "=" + landed_expr(mode, m, xcol, "B") + f"-$B${r}", fmt=CUR2)
            # Freight effect: scenario landed minus (scenario FX, plan rate)
            w(ws, f"{FR}{r}", f"={L}{r}-(" + landed_expr(mode, m, xcol, "B") + ")", fmt=CUR2)
            # Total effect
            w(ws, f"{TT}{r}", f"={L}{r}-$B${r}", fmt=CUR2, bold=True)
    # average row
    ar = start_row + 7
    w(ws, f"A{ar}", "Average", bold=True, italic=True)
    for col in ["B"] + [get_column_letter(3 + k) for k in range(12)]:
        w(ws, f"{col}{ar}", f"=AVERAGE({col}{start_row+1}:{col}{start_row+6})",
          fmt=CUR2, bold=True, italic=True)
    return ar


section(s2, 13, "AIR FREIGHT — Landed Cost per Unit (AUD)", "N")
air_avg = landed_block(s2, 14, "air", 6)          # SKUs at rows 15-20

section(s2, 24, "SEA FREIGHT — Landed Cost per Unit (AUD)", "N")
sea_avg = landed_block(s2, 25, "sea", 6)          # SKUs at rows 26-31

w(s2, "A34", "Read it like this:  under Severe, compare each SKU's FX Δ against its Freight Δ to see which "
             "shock is doing the damage. Light, high-value SKUs are FX-led; bulky, heavy SKUs are freight-led.",
  italic=True, size=9, color="595959", wrap=True)
s2.merge_cells("A34:N35")

# red for negative Freight Δ / positive cost increases we leave black; here Total Δ up = cost up (bad) -> red
from openpyxl.utils import get_column_letter
for blk in (14, 25):
    for j in range(3):
        tcol = get_column_letter(3 + j * 4 + 3)   # Total Δ column of each scenario
        rng = f"{tcol}{blk+1}:{tcol}{blk+7}"
        s2.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=["0"],
                            font=Font(name=FONT, color="9C0006"),
                            fill=PatternFill("solid", fgColor="FFC7CE")))

for col, wd in zip("ABCDEFGHIJKLMN",
                   (10, 20, 10, 9, 10, 10, 10, 9, 10, 10, 10, 9, 10, 10)):
    s2.column_dimensions[col].width = wd
s2.freeze_panes = "B15"

# =========================================================================
# TAB 3 - REVENUE SCENARIO
# =========================================================================
s3 = wb.create_sheet("3. Revenue Scenario")
s3.sheet_view.showGridLines = False
title(s3, "A1", "Revenue Scenario  —  Monthly by Segment, Elasticity-Driven", "H1")
w(s3, "A2", "Scenario revenue = Plan × (1 + consumer-spending move × segment elasticity). Price is held constant, "
             "so the move is pure volume. Variance columns are scenario minus plan (negatives in red).",
  italic=True, size=9, color="595959", wrap=True)
s3.merge_cells("A2:H2")
s3.row_dimensions[2].height = 26

months = ["Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026", "Jan 2027",
          "Feb 2027", "Mar 2027", "Apr 2027", "May 2027", "Jun 2027"]
season = [1.00, 1.05, 1.30, 1.35, 0.80, 0.85, 0.95, 1.00, 1.05, 1.10]
seg_defs = [("PREMIUM", 180000, 17), ("CORE", 320000, 18), ("VALUE", 150000, 19)]

HDR3 = ["Month", "Plan (AUD)", "Base", "Moderate", "Severe",
        "Var Base", "Var Moderate", "Var Severe"]

seg_first_rows = {}   # segment name -> first month row (for the Total block)
row = 4
for name, base, elast_row in seg_defs:
    section(s3, row, f"{name} SEGMENT", "H")
    colhdr(s3, row + 1, HDR3)
    first = row + 2
    seg_first_rows[name] = first
    for k, mth in enumerate(months):
        r = first + k
        plan_val = round(base * season[k], -2)
        w(s3, f"A{r}", mth)
        w(s3, f"B{r}", plan_val, color=BLUE, fmt=CUR0)
        # scenario cols reference plan cell * (1 + spending * elasticity)
        for col, scol in zip("CDE", "CDE"):
            w(s3, f"{col}{r}", f"=$B{r}*(1+{AS}${scol}$12*{AS}$B${elast_row})", fmt=CUR0)
        w(s3, f"F{r}", f"=C{r}-$B{r}", fmt=CUR0)
        w(s3, f"G{r}", f"=D{r}-$B{r}", fmt=CUR0)
        w(s3, f"H{r}", f"=E{r}-$B{r}", fmt=CUR0)
    sub = first + 10
    w(s3, f"A{sub}", f"{name} subtotal", bold=True)
    for col in "BCDEFGH":
        w(s3, f"{col}{sub}", f"=SUM({col}{first}:{col}{first+9})", fmt=CUR0, bold=True,
          fill="F2F2F2")
    row = sub + 2

# TOTAL block
section(s3, row, "TOTAL — ALL SEGMENTS", "H")
colhdr(s3, row + 1, HDR3)
tot_first = row + 2
pf, cf, vf = seg_first_rows["PREMIUM"], seg_first_rows["CORE"], seg_first_rows["VALUE"]
for k, mth in enumerate(months):
    r = tot_first + k
    w(s3, f"A{r}", mth)
    for col in "BCDEFGH":
        w(s3, f"{col}{r}", f"={col}{pf+k}+{col}{cf+k}+{col}{vf+k}", fmt=CUR0,
          bold=(col in "BCDE"))
tot_sub = tot_first + 10       # <-- period totals row referenced by Tab 4
w(s3, f"A{tot_sub}", "PERIOD TOTAL", bold=True, color="FFFFFF", fill=NAVY)
for col in "BCDEFGH":
    w(s3, f"{col}{tot_sub}", f"=SUM({col}{tot_first}:{col}{tot_first+9})",
      fmt=CUR0, bold=True, color="FFFFFF", fill=NAVY)

# store for Tab 4
REV_PLAN_TOTAL = f"$B${tot_sub}"
REV_BASE_TOTAL = f"$C${tot_sub}"
REV_MOD_TOTAL = f"$D${tot_sub}"
REV_SEV_TOTAL = f"$E${tot_sub}"

# conditional format negative variances red across all variance columns
for r0 in [seg_first_rows["PREMIUM"], seg_first_rows["CORE"], seg_first_rows["VALUE"], tot_first]:
    s3.conditional_formatting.add(
        f"F{r0}:H{r0+10}",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(name=FONT, color="9C0006"),
                   fill=PatternFill("solid", fgColor="FFC7CE")))

for col, wd in zip("ABCDEFGH", (16, 14, 14, 14, 14, 13, 13, 13)):
    s3.column_dimensions[col].width = wd

# =========================================================================
# TAB 4 - MARGIN BRIDGE
# =========================================================================
s4 = wb.create_sheet("4. Margin Bridge")
s4.sheet_view.showGridLines = False
title(s4, "A1", "Margin Bridge  —  Planned Gross Profit to Scenario Gross Profit", "G1")
w(s4, "A2", "Period gross profit (AUD), decomposed into FX, air-freight, sea-freight substitution and demand-volume "
             "steps. The four steps sum exactly to the gap between planned and scenario GP.",
  italic=True, size=9, color="595959", wrap=True)
s4.merge_cells("A2:G2")
s4.row_dimensions[2].height = 26

# --- Plan economics ---
section(s4, 4, "PLAN ECONOMICS  (period)", "G")
colhdr(s4, 5, ["Item", "Value", "", "", "", "", "Notes"])
plan_econ = [
    (6,  "Planned units",            f"={AS}$B$23", NUM0, GREEN, "from Assumptions"),
    (7,  "Planned ASP (AUD)",        f"={AS}$B$24", CUR2, GREEN, "from Assumptions"),
    (8,  "Planned revenue",          "=B6*B7", CUR0, BLACK, "units × ASP"),
    (9,  "Product cost / unit",      f"={AS}$B$25*(1+{AS}$B$10+{AS}$B$11)*{AS}$B$7", CUR2, BLACK, "FOB × (1+duty+ins) × AUD/USD"),
    (10, "Air freight / unit",       f"={AS}$B$26*{AS}$B$8*{AS}$B$7", CUR2, BLACK, "weight × air rate × AUD/USD"),
    (11, "Sea freight / unit",       f"=(1/{AS}$B$27)*{AS}$B$9*{AS}$B$7", CUR2, BLACK, "(1/units-per-CBM) × sea rate × AUD/USD"),
    (12, "Air share (plan)",         f"={AS}$B$28", PCT1, GREEN, "from Assumptions"),
    (13, "Sea share (plan)",         "=1-B12", PCT1, BLACK, "balance"),
    (14, "Blended freight / unit",   "=B12*B10+B13*B11", CUR2, BLACK, "mix-weighted"),
    (15, "COGS / unit (plan)",       "=B9+B14", CUR2, BLACK, "product + freight"),
    (16, "Planned COGS",             "=B6*B15", CUR0, BLACK, "units × COGS/unit"),
    (17, "Planned gross profit",     "=B8-B16", CUR0, BLACK, "revenue − COGS"),
    (18, "Planned gross margin %",   "=B17/B8", PCT1, BLACK, "GP / revenue"),
]
for r, label, formula, fmt, color, note in plan_econ:
    w(s4, f"A{r}", label)
    w(s4, f"B{r}", formula, fmt=fmt, color=color, bold=(r in (8, 17, 18)))
    w(s4, f"G{r}", note, size=9, color="595959")

# --- Scenario bridge (scenarios in columns C/D/E) ---
section(s4, 20, "SCENARIO BRIDGE  (period, AUD)", "G")
w(s4, "A21", "Bridge line", bold=True, fill=LIGHTBLUE, border=HDR_BORDER)
w(s4, "B21", "", fill=LIGHTBLUE, border=HDR_BORDER)
for col, label in zip("CDE", ["Base", "Moderate", "Severe"]):
    w(s4, f"{col}21", label, bold=True, align="center", fill=LIGHTBLUE, border=HDR_BORDER)
w(s4, "G21", "Notes", bold=True, fill=LIGHTBLUE, border=HDR_BORDER)
s4.row_dimensions[21].height = 16

# per-scenario formula rows. xcol maps Base->C, Moderate->D, Severe->E in Assumptions
bridge_rows = [
    (22, "FX ratio (scen ÷ plan AUD/USD)", lambda c: f"={AS}${c}$7/{AS}$B$7", MULT, "revaluation factor on USD costs"),
    (23, "Air rate ratio",                 lambda c: f"={AS}${c}$8/{AS}$B$8", MULT, ""),
    (24, "Sea rate ratio",                 lambda c: f"={AS}${c}$9/{AS}$B$9", MULT, ""),
    (25, "Air share (scenario)",           lambda c: f"=$B$12-{AS}${c}$13", PCT1, "plan air share less substitution shift"),
    (26, "Sea share (scenario)",           lambda c, rr=26: f"=1-{c}25", PCT1, ""),
    (27, "Demand factor (scen ÷ plan rev)",None, MULT, "from Revenue Scenario period totals"),
    (28, "c0  COGS/unit @ plan",           lambda c: "=$B$15", CUR2, ""),
    (29, "Product/unit @ scenario",        lambda c: f"=$B$9*{c}22", CUR2, "plan × FX ratio"),
    (30, "Air freight/unit @ scenario",    lambda c: f"=$B$10*{c}22*{c}23", CUR2, "plan × FX × air-rate ratio"),
    (31, "Sea freight/unit @ scenario",    lambda c: f"=$B$11*{c}22*{c}24", CUR2, "plan × FX × sea-rate ratio"),
    (32, "c1  after FX",                    lambda c: f"={c}28*{c}22", CUR2, ""),
    (33, "c2  after air rate",              lambda c: f"={c}29+$B$12*{c}30+$B$13*($B$11*{c}22)", CUR2, ""),
    (34, "c3  after sea sub (=scen COGS/u)",lambda c: f"={c}29+{c}25*{c}30+{c}26*{c}31", CUR2, ""),
]
# demand factor references revenue totals per scenario column
demand_ref = {"C": f"={RS}{REV_BASE_TOTAL}/{RS}{REV_PLAN_TOTAL}",
              "D": f"={RS}{REV_MOD_TOTAL}/{RS}{REV_PLAN_TOTAL}",
              "E": f"={RS}{REV_SEV_TOTAL}/{RS}{REV_PLAN_TOTAL}"}
for r, label, fn, fmt, note in bridge_rows:
    w(s4, f"A{r}", label, size=9, italic=(r not in ()))
    for c in "CDE":
        if r == 27:
            w(s4, f"{c}{r}", demand_ref[c], fmt=fmt, align="center")
        else:
            w(s4, f"{c}{r}", fn(c), fmt=fmt, align="center")
    w(s4, f"G{r}", note, size=8, color="595959")

# the bridge step rows (the deliverable)
step_rows = [
    (35, "STEP — FX effect",                lambda c: f"=-$B$6*({c}32-{c}28)"),
    (36, "STEP — Air freight",              lambda c: f"=-$B$6*({c}33-{c}32)"),
    (37, "STEP — Sea freight substitution", lambda c: f"=-$B$6*({c}34-{c}33)"),
    (38, "STEP — Demand volume",            lambda c: f"=({c}27-1)*($B$8-$B$6*{c}34)"),
]
for r, label, fn in step_rows:
    w(s4, f"A{r}", label, bold=True)
    for c in "CDE":
        w(s4, f"{c}{r}", fn(c), fmt=CUR0, align="center")
# results
w(s4, "A39", "Scenario gross profit", bold=True, fill="F2F2F2")
w(s4, "B39", "Plan:", italic=True, size=9, align="right")
for c in "CDE":
    w(s4, f"{c}39", f"=$B$17+{c}35+{c}36+{c}37+{c}38", fmt=CUR0, bold=True, fill="F2F2F2", align="center")
w(s4, "A40", "Scenario revenue")
for c in "CDE":
    w(s4, f"{c}40", f"=$B$8*{c}27", fmt=CUR0, align="center")
w(s4, "A41", "Scenario COGS")
for c in "CDE":
    w(s4, f"{c}41", f"=$B$6*{c}27*{c}34", fmt=CUR0, align="center")
w(s4, "A42", "Scenario gross margin %", bold=True)
for c in "CDE":
    w(s4, f"{c}42", f"={c}39/{c}40", fmt=PCT1, bold=True, align="center")
w(s4, "A43", "Scenario units (for levers)", size=9, italic=True)
for c in "CDE":
    w(s4, f"{c}43", f"=$B$6*{c}27", fmt=NUM0, align="center")
w(s4, "A44", "FX cost @ scenario volume", size=9, italic=True)
for c in "CDE":
    w(s4, f"{c}44", f"={c}43*{c}28*({c}22-1)", fmt=CUR0, align="center")
w(s4, "A45", "Check: GP − (Rev−COGS) = 0", size=8, italic=True, color="595959")
for c in "CDE":
    w(s4, f"{c}45", f"={c}39-({c}40-{c}41)", fmt=CUR0, align="center", size=8)
w(s4, "B39", "Plan GP →", italic=True, size=9, align="right", color="595959")

# red for negative step values
s4.conditional_formatting.add("C35:E38", CellIsRule(
    operator="lessThan", formula=["0"], font=Font(name=FONT, color="9C0006")))

# --- Waterfall data (Severe = column E) + chart ---
section(s4, 48, "WATERFALL DATA — Severe scenario (period gross profit, AUD)", "G")
colhdr(s4, 49, ["Step", "Value", "Cumulative", "Base", "Decrease", "Increase", "Total"])
# Plan GP (total bar)
w(s4, "A50", "Planned GP")
w(s4, "B50", "=$B$17", fmt=CUR0)
w(s4, "C50", "=B50", fmt=CUR0)
w(s4, "D50", 0, fmt=CUR0); w(s4, "E50", 0, fmt=CUR0); w(s4, "F50", 0, fmt=CUR0)
w(s4, "G50", "=B50", fmt=CUR0)
wf = [("A51", "FX effect", "=$E$35"),
      ("A52", "Air freight", "=$E$36"),
      ("A53", "Sea freight substitution", "=$E$37"),
      ("A54", "Demand volume", "=$E$38")]
prev = 50
for coord, label, val in wf:
    r = int(coord[1:])
    w(s4, f"A{r}", label)
    w(s4, f"B{r}", val, fmt=CUR0)
    w(s4, f"C{r}", f"=C{prev}+B{r}", fmt=CUR0)          # cumulative
    w(s4, f"E{r}", f"=MAX(-B{r},0)", fmt=CUR0)          # decrease height
    w(s4, f"F{r}", f"=MAX(B{r},0)", fmt=CUR0)           # increase height
    w(s4, f"D{r}", f"=C{prev}-E{r}", fmt=CUR0)          # invisible base
    w(s4, f"G{r}", 0, fmt=CUR0)
    prev = r
# Scenario GP (total bar)
w(s4, "A55", "Scenario GP")
w(s4, "B55", "=$E$39", fmt=CUR0)
w(s4, "C55", "=B55", fmt=CUR0)
w(s4, "D55", 0, fmt=CUR0); w(s4, "E55", 0, fmt=CUR0); w(s4, "F55", 0, fmt=CUR0)
w(s4, "G55", "=B55", fmt=CUR0)

# waterfall chart (stacked bar, invisible base)
import os as _os
_SKIP_CHART = _os.environ.get("SKIP_CHART") == "1"
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
chart.overlap = 100
chart.title = "Planned GP → Severe Scenario GP"
chart.y_axis.title = "AUD"
chart.height = 8.5
chart.width = 18
cats = Reference(s4, min_col=1, min_row=50, max_row=55)
for col_idx, name, rgb, hide in [(4, "Base", None, True), (5, "Decrease", "C0392B", False),
                                  (6, "Increase", "27AE60", False), (7, "Total", "2E5AAC", False)]:
    ref = Reference(s4, min_col=col_idx, min_row=49, max_row=55)
    ser = Series(ref, title_from_data=True)
    if hide:
        ser.graphicalProperties.noFill = True
        ser.graphicalProperties.line.noFill = True
    else:
        ser.graphicalProperties.solidFill = rgb
    chart.series.append(ser)
chart.set_categories(cats)
chart.legend.position = "b"
if not _SKIP_CHART:
    s4.add_chart(chart, "I4")

for col, wd in zip("ABCDEFG", (30, 14, 14, 12, 12, 12, 14)):
    s4.column_dimensions[col].width = wd

# =========================================================================
# TAB 5 - MITIGATION LEVERS
# =========================================================================
s5 = wb.create_sheet("5. Mitigation Levers")
s5.sheet_view.showGridLines = False
title(s5, "A1", "Mitigation Levers  —  Closing the Gap", "F1")
w(s5, "A2", "Each lever is modelled against the selected scenario's gross-profit gap. Change the yellow scenario "
             "selector and the blue lever settings (X, Y, Z); the residual updates so you can find the combination "
             "that closes the gap.", italic=True, size=9, color="595959", wrap=True)
s5.merge_cells("A2:F2")
s5.row_dimensions[2].height = 30

# scenario selector
section(s5, 4, "SCENARIO", "F")
w(s5, "A5", "Scenario:", bold=True)
w(s5, "B5", "Severe", color=BLUE, bold=True, fill=YELLOW, align="center", border=BOX)
dv = DataValidation(type="list", formula1='"Base,Moderate,Severe"', allow_blank=False)
s5.add_data_validation(dv)
dv.add(s5["B5"])
w(s5, "C5", "← pick Base / Moderate / Severe", italic=True, size=9, color="595959")
w(s5, "A6", "Scenario column index", size=9, italic=True, color="595959")
w(s5, "B6", f"=MATCH($B$5,{MB}$C$21:$E$21,0)", align="center", size=9, color="595959")

# pull scenario values via INDEX(range,1,index)
def idx(rng):
    return f"=INDEX({MB}{rng},1,$B$6)"

pulls = [
    (7,  "Scenario gross profit",      idx("$C$39:$E$39"), CUR0),
    (8,  "Planned gross profit",       f"={MB}$B$17", CUR0),
    (9,  "Gap to plan (scenario − plan)", "=B7-B8", CUR0),
    (10, "Scenario units",             idx("$C$43:$E$43"), NUM0),
    (11, "Air freight / unit (scen)",  idx("$C$30:$E$30"), CUR2),
    (12, "Sea freight / unit (scen)",  idx("$C$31:$E$31"), CUR2),
    (13, "FX cost @ scenario volume",  idx("$C$44:$E$44"), CUR0),
    (14, "Scenario revenue",           idx("$C$40:$E$40"), CUR0),
    (15, "Scenario COGS",              idx("$C$41:$E$41"), CUR0),
    (16, "Price elasticity",           f"={AS}$B$29", "0.00"),
    (17, "Scenario air share (cap for X)", idx("$C$25:$E$25"), PCT1),
]
for r, label, formula, fmt in pulls:
    w(s5, f"A{r}", label, italic=(r not in (7, 8, 9)))
    w(s5, f"B{r}", formula, fmt=fmt, color=(GREEN if r in (7, 8, 10, 11, 12, 13, 14, 15, 16, 17) else BLACK),
      bold=(r == 9))
w(s5, "D9", "the gap the levers must close", italic=True, size=9, color="595959")

# levers
section(s5, 18, "MITIGATION LEVERS", "F")
w(s5, "A19", "Lever", bold=True, fill=LIGHTBLUE, border=HDR_BORDER)
w(s5, "B19", "Setting", bold=True, fill=LIGHTBLUE, align="center", border=HDR_BORDER)
w(s5, "C19", "Modelled GP impact (AUD)", bold=True, fill=LIGHTBLUE, align="center", border=HDR_BORDER)
w(s5, "D19", "How it is modelled", bold=True, fill=LIGHTBLUE, border=HDR_BORDER)
s5.row_dimensions[19].height = 16

# Lever A - shift to sea
w(s5, "A20", "A.  Shift volume to sea")
w(s5, "B20", 0.12, color=BLUE, fmt=PCT1, align="center", fill=YELLOW, border=BOX)
w(s5, "C20", "=B20*$B$10*($B$11-$B$12)", fmt=CUR0, bold=True)
w(s5, "D20", "X% of scenario units moved air→sea × (air − sea) freight/unit. Keep X ≤ scenario air share (B17).",
  size=9, color="595959", wrap=True)
# Lever B - hedge
w(s5, "A21", "B.  Forward-hedge USD exposure")
w(s5, "B21", 0.75, color=BLUE, fmt=PCT1, align="center", fill=YELLOW, border=BOX)
w(s5, "C21", "=B21*$B$13", fmt=CUR0, bold=True)
w(s5, "D21", "Y% of the scenario FX cost locked at the plan rate (hedge cost ignored).",
  size=9, color="595959", wrap=True)
# Lever C - price passthrough
w(s5, "A22", "C.  Pass through price increase")
w(s5, "B22", 0.06, color=BLUE, fmt=PCT1, align="center", fill=YELLOW, border=BOX)
w(s5, "C22", "=$B$14*((1+B22)*(1-B22*$B$16)-1)+$B$15*B22*$B$16", fmt=CUR0, bold=True)
w(s5, "D22", "Z% price rise on scenario revenue, less the volume lost at price elasticity; COGS falls with volume.",
  size=9, color="595959", wrap=True)

# combined + residual
w(s5, "A24", "Combined lever impact", bold=True, fill="F2F2F2")
w(s5, "C24", "=SUM(C20:C22)", fmt=CUR0, bold=True, fill="F2F2F2")
w(s5, "A25", "Scenario GP after levers", bold=True)
w(s5, "C25", "=B7+C24", fmt=CUR0, bold=True)
w(s5, "A26", "Residual vs plan", bold=True)
w(s5, "C26", "=C25-B8", fmt=CUR0, bold=True)
w(s5, "A27", "Gap closed?", bold=True)
w(s5, "C27", '=IF(C26>=0,"YES — gap closed","NO — shortfall remaining")', bold=True, align="left")

# red/green formatting
s5.conditional_formatting.add("B9", CellIsRule(operator="lessThan", formula=["0"],
                              font=Font(name=FONT, bold=True, color="9C0006")))
for cell in ("C26",):
    s5.conditional_formatting.add(cell, CellIsRule(operator="lessThan", formula=["0"],
                                  font=Font(name=FONT, bold=True, color="9C0006"),
                                  fill=PatternFill("solid", fgColor="FFC7CE")))
    s5.conditional_formatting.add(cell, CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                                  font=Font(name=FONT, bold=True, color="006100"),
                                  fill=PatternFill("solid", fgColor="C6EFCE")))
s5.conditional_formatting.add("C27", FormulaRule(formula=['ISNUMBER(SEARCH("NO",C27))'],
                              font=Font(name=FONT, bold=True, color="9C0006")))
s5.conditional_formatting.add("C27", FormulaRule(formula=['ISNUMBER(SEARCH("YES",C27))'],
                              font=Font(name=FONT, bold=True, color="006100")))

w(s5, "A29", "Tip: levers are shown standalone (second-order interactions are small and ignored). "
             "Combine them until Residual vs plan turns green.", italic=True, size=9, color="595959", wrap=True)
s5.merge_cells("A29:F30")

for col, wd in zip("ABCDEF", (30, 12, 22, 40, 4, 4)):
    s5.column_dimensions[col].width = wd

# force Excel / Sheets to recalculate every formula on open (no stale/blank cells)
wb.calculation.fullCalcOnLoad = True

wb.save(OUT)
print("wrote", OUT)
